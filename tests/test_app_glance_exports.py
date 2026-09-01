from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from india_resilience_tool.app.glance_exports import (
    DRIVER_UNAVAILABLE_NOTE,
    build_glance_answer_pack_xlsx,
    build_glance_answer_text,
    build_glance_csv_bytes,
    build_glance_export_frame,
    glance_export_filename,
)
from india_resilience_tool.app.landing_runtime import _compute_visible_ranking_rows


def _district_scores() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "state_name": ["Telangana", "Telangana", "Maharashtra"],
            "district_name": ["Medchal-Malkajgiri", "Nalgonda", "Pune"],
            "__state_key": ["telangana", "telangana", "maharashtra"],
            "__district_key": ["telangana|medchal_malkajgiri", "telangana|nalgonda", "maharashtra|pune"],
            "bundle_score": [92.0, 81.0, 74.0],
            "bundle_score_display": ["92.0", "81.0", "74.0"],
            "score_band": ["High", "Moderate", "Low"],
            "district_rank": [1, 2, 1],
            "district_count": [2, 2, 1],
        }
    )


def _block_scores() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "state_name": ["Telangana", "Telangana", "Telangana"],
            "district_name": ["Nalgonda", "Nalgonda", "Khammam"],
            "block_name": ["Chityal", "Narketpalle", "Wyra"],
            "__state_key": ["telangana", "telangana", "telangana"],
            "__district_key": ["telangana|nalgonda", "telangana|nalgonda", "telangana|khammam"],
            "__block_key": ["telangana|nalgonda|chityal", "telangana|nalgonda|narketpalle", "telangana|khammam|wyra"],
            "bundle_score": [91.0, 62.0, 44.0],
            "bundle_score_display": ["91.0", "62.0", "44.0"],
            "score_band": ["High", "Moderate", "Low"],
            "block_rank_within_district": [1, 2, 1],
            "block_count_within_district": [2, 2, 1],
        }
    )


def _drivers() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "scope_level": ["district", "district", "district", "district", "district", "district", "block"],
            "state_name": ["Telangana", "Telangana", "Telangana", "Telangana", "Telangana", "Telangana", "Telangana"],
            "district_name": [
                "Medchal-Malkajgiri",
                "Medchal-Malkajgiri",
                "Medchal-Malkajgiri",
                "Nalgonda",
                "Nalgonda",
                "Nalgonda",
                "Nalgonda",
            ],
            "block_name": ["", "", "", "", "", "", "Chityal"],
            "__district_key": [
                "telangana|medchal–malkajgiri",
                "telangana|medchal–malkajgiri",
                "telangana|medchal–malkajgiri",
                "telangana|nalgonda",
                "telangana|nalgonda",
                "telangana|nalgonda",
                "telangana|nalgonda",
            ],
            "__block_key": ["", "", "", "", "", "", "telangana|nalgonda|chityal"],
            "driver_rank": [1, 2, 3, 1, 2, 3, 1],
            "driver_label": [
                "Hot days",
                "Warm nights",
                "Humidity",
                "Rainfall",
                "Runoff",
                "Saturation",
                "Local heat",
            ],
            "driver_score": [0.9, 0.8, 0.7, 0.5, 0.4, 0.3, 0.6],
        }
    )


def test_district_export_filters_state_uses_persisted_rank_and_names_top_districts() -> None:
    visible = _compute_visible_ranking_rows(
        focus_level="state",
        selected_state="Telangana",
        selected_district=None,
        selected_block=None,
        state_scores=pd.DataFrame(),
        district_scores=_district_scores(),
    )
    export, note = build_glance_export_frame(visible, _drivers())
    answer = build_glance_answer_text(
        export,
        bundle_label="Heat Risk",
        scenario_label="SSP5-8.5",
        period_label="2040-2060",
        geography_label="Telangana",
        is_projection=True,
        driver_note=note,
    )
    assert export["unit_name"].tolist() == ["Medchal-Malkajgiri", "Nalgonda"]
    assert export["rank"].tolist() == [1, 2]
    assert export.loc[0, "top_driver_1"] == "Hot days"
    assert "Medchal-Malkajgiri, Nalgonda" in answer


def test_district_focus_answer_leads_with_current_focus_and_then_visible_top_three() -> None:
    visible = _compute_visible_ranking_rows(
        focus_level="district",
        selected_state="Telangana",
        selected_district="Nalgonda",
        selected_block=None,
        state_scores=pd.DataFrame(),
        district_scores=_district_scores(),
    )
    export, note = build_glance_export_frame(visible, _drivers())
    answer = build_glance_answer_text(
        export,
        bundle_label="Extreme Rainfall | Flash Flood Risk",
        scenario_label="SSP5-8.5",
        period_label="Mid-century",
        geography_label="Telangana",
        is_projection=True,
        driver_note=note,
    )

    assert answer.startswith("For Telangana district rankings, Nalgonda is the current focus.")
    assert "It ranks 2 / 2" in answer
    assert "with a bundle score of 81.0 and Moderate risk band" in answer
    assert "Its leading drivers are Rainfall, Runoff, Saturation." in answer
    assert "The highest-ranked districts in the visible table are Medchal-Malkajgiri, Nalgonda." in answer


def test_unfocused_answer_preserves_top_visible_ranking_wording() -> None:
    visible = _compute_visible_ranking_rows(
        focus_level="state",
        selected_state="Telangana",
        selected_district=None,
        selected_block=None,
        state_scores=pd.DataFrame(),
        district_scores=_district_scores(),
    )
    export, note = build_glance_export_frame(visible, _drivers())
    answer = build_glance_answer_text(
        export,
        bundle_label="Heat Risk",
        scenario_label="SSP5-8.5",
        period_label="2040-2060",
        geography_label="Telangana",
        is_projection=True,
        driver_note=note,
    )

    assert answer.startswith("For Telangana, the visible Glance ranking")
    assert "current focus" not in answer


def test_block_focus_answer_leads_with_selected_block() -> None:
    visible = _compute_visible_ranking_rows(
        focus_level="block",
        selected_state="Telangana",
        selected_district="Nalgonda",
        selected_block="Chityal",
        state_scores=pd.DataFrame(),
        district_scores=pd.DataFrame(),
        block_scores=_block_scores(),
    )
    export, note = build_glance_export_frame(visible, _drivers())
    answer = build_glance_answer_text(
        export,
        bundle_label="Heat Risk",
        scenario_label="SSP5-8.5",
        period_label="Mid-century",
        geography_label="Nalgonda",
        is_projection=True,
        driver_note=note,
    )

    assert answer.startswith("For Nalgonda block rankings, Chityal is the current focus.")
    assert "It ranks 1 / 2" in answer
    assert "Its leading drivers are Local heat." in answer


def test_driver_join_survives_punctuation_dash_variants() -> None:
    visible = _compute_visible_ranking_rows(
        focus_level="state",
        selected_state="Telangana",
        selected_district=None,
        selected_block=None,
        state_scores=pd.DataFrame(),
        district_scores=_district_scores().drop(columns=["__district_key"]),
    )
    export, _note = build_glance_export_frame(visible, _drivers())
    assert export.loc[0, "top_driver_2"] == "Warm nights"


def test_block_export_filters_selected_district_and_marks_current_focus() -> None:
    visible = _compute_visible_ranking_rows(
        focus_level="block",
        selected_state="Telangana",
        selected_district="Nalgonda",
        selected_block="Chityal",
        state_scores=pd.DataFrame(),
        district_scores=pd.DataFrame(),
        block_scores=_block_scores(),
    )
    export, _note = build_glance_export_frame(visible, _drivers())
    assert export["unit_name"].tolist() == ["Chityal", "Narketpalle"]
    assert export["rank"].tolist() == [1, 2]
    assert export["is_current_focus"].tolist() == [True, False]


def test_missing_drivers_still_export_and_records_note() -> None:
    visible = _compute_visible_ranking_rows(
        focus_level="state",
        selected_state="Telangana",
        selected_district=None,
        selected_block=None,
        state_scores=pd.DataFrame(),
        district_scores=_district_scores(),
    )
    export, note = build_glance_export_frame(visible, pd.DataFrame())
    assert note == DRIVER_UNAVAILABLE_NOTE
    assert export["top_driver_1"].fillna("").tolist() == ["", ""]


def test_empty_visible_rows_return_empty_export_frame() -> None:
    visible = _compute_visible_ranking_rows(
        focus_level="state",
        selected_state="Kerala",
        selected_district=None,
        selected_block=None,
        state_scores=pd.DataFrame(),
        district_scores=_district_scores(),
    )
    export, note = build_glance_export_frame(visible, _drivers())
    assert export.empty
    assert note == ""


def test_snapshot_branch_uses_non_projection_wording() -> None:
    visible = _compute_visible_ranking_rows(
        focus_level="state",
        selected_state="Telangana",
        selected_district=None,
        selected_block=None,
        state_scores=pd.DataFrame(),
        district_scores=_district_scores(),
    )
    export, note = build_glance_export_frame(visible, _drivers())
    answer = build_glance_answer_text(
        export,
        bundle_label="Riverine Flood",
        scenario_label="snapshot",
        period_label="Current",
        geography_label="Telangana",
        is_projection=False,
        driver_note=note,
    )
    assert "for Current" in answer


def test_csv_bytes_are_utf8_sig_and_include_expected_columns() -> None:
    visible = _compute_visible_ranking_rows(
        focus_level="state",
        selected_state="Telangana",
        selected_district=None,
        selected_block=None,
        state_scores=pd.DataFrame(),
        district_scores=_district_scores(),
    )
    export, _note = build_glance_export_frame(visible, _drivers())
    csv_bytes = build_glance_csv_bytes(export)
    assert csv_bytes.startswith(b"\xef\xbb\xbf")
    decoded = csv_bytes.decode("utf-8-sig")
    assert "unit_name" in decoded
    assert "is_current_focus" in decoded
    assert "Medchal-Malkajgiri" in decoded


def test_csv_bytes_mark_selected_district_as_current_focus() -> None:
    visible = _compute_visible_ranking_rows(
        focus_level="district",
        selected_state="Telangana",
        selected_district="Nalgonda",
        selected_block=None,
        state_scores=pd.DataFrame(),
        district_scores=_district_scores(),
    )
    export, _note = build_glance_export_frame(visible, _drivers())
    decoded = build_glance_csv_bytes(export).decode("utf-8-sig")
    assert "is_current_focus" in decoded
    assert "Nalgonda" in decoded
    assert "True" in decoded


def test_xlsx_contains_expected_sheets() -> None:
    visible = _compute_visible_ranking_rows(
        focus_level="state",
        selected_state="Telangana",
        selected_district=None,
        selected_block=None,
        state_scores=pd.DataFrame(),
        district_scores=_district_scores(),
    )
    export, note = build_glance_export_frame(visible, _drivers())
    xlsx = build_glance_answer_pack_xlsx(
        answer_text="Answer",
        export_frame=export,
        metadata={"driver_source_artifact": "drivers.parquet"},
        driver_note=note,
    )
    workbook = load_workbook(BytesIO(xlsx), read_only=True)
    assert workbook.sheetnames == ["Answer", "Ranking", "Drivers", "Metadata", "Method Notes"]


def test_xlsx_ranking_sheet_includes_current_focus_column() -> None:
    visible = _compute_visible_ranking_rows(
        focus_level="district",
        selected_state="Telangana",
        selected_district="Nalgonda",
        selected_block=None,
        state_scores=pd.DataFrame(),
        district_scores=_district_scores(),
    )
    export, note = build_glance_export_frame(visible, _drivers())
    xlsx = build_glance_answer_pack_xlsx(
        answer_text="Answer",
        export_frame=export,
        metadata={"driver_source_artifact": "drivers.parquet"},
        driver_note=note,
    )
    workbook = load_workbook(BytesIO(xlsx), read_only=True)
    ranking_sheet = workbook["Ranking"]
    headers = [cell.value for cell in next(ranking_sheet.iter_rows(min_row=1, max_row=1))]
    assert "is_current_focus" in headers


def test_filename_sanitizer_handles_spaces_slashes_punctuation_and_band() -> None:
    filename = glance_export_filename(
        kind="xlsx",
        bundle_slug="Flood & Extreme/Rainfall Risk",
        unit_scope="district",
        scenario="snapshot",
        period="Current",
        geography="Medchal / Malkajgiri!",
        band_filter="Very High",
    )
    assert filename == (
        "irt_glance_answer_pack_flood_extreme_rainfall_risk_district_snapshot_current_"
        "medchal_malkajgiri_very_high.xlsx"
    )
